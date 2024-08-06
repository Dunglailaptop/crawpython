from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from concurrent.futures import ThreadPoolExecutor, as_completed
from unidecode import unidecode 
from tkinter import ttk, filedialog, Tk
from datetime import datetime
from tkinter import *
from tkcalendar import DateEntry
from tkinter.filedialog import askopenfilename
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import pandas as pd
import openpyxl
import pandas as pd
import math
import time
import csv
import json
import unicodedata
import re
import requests
import io
import os
from PIL import Image
import numpy as np

dateSelect = ""
page = 0
#ghi file excel từ csv vào đây
urlFolder = ""
#ghi file csv 
urlCsvFile = ""
#tiêu đề bảng
dataheader = []
#chọn ngày
def select_date(driver, month, year):
    try:
        # Chọn năm
        year_select = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'ui-datepicker-year'))
        )
        year_select.click()
        year_option = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//option[@value='{year}']"))
        )
        year_option.click()

        # Chọn tháng
        month_select = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'ui-datepicker-month'))
        )
        month_select.click()
        month_option = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//option[@value='{month}']"))
        )
        month_option.click()

    except Exception as e:
        print(f"Lỗi khi chọn tháng và năm: {e}")

def set_date2(driver, element_id, date_value):
    try:
        # Chờ đến khi phần tử có ID xuất hiện
        searchIdDateTime = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, element_id))
        )

        # Tìm thẻ <span> với class 'input-group-addon' bên trong phần tử
        findspan = searchIdDateTime.find_element(By.CLASS_NAME, 'input-group-addon')

        # Click vào thẻ <span> để mở lịch
        findspan.click()
        time.sleep(2)  # Thêm thời gian chờ để đảm bảo lịch được mở

        # Tách date_value thành ngày, tháng và năm
        day, month, year = date_value.split('/')
        day = int(day)
        month = int(month) - 1  # Chuyển đổi tháng về dạng số và trừ 1 vì tháng trong datepicker bắt đầu từ 0
        year = int(year)

        # Gọi hàm select_date để chọn tháng và năm
        select_date(driver, month, year)

        # Chọn ngày
        day_xpath = f"//td[not(contains(@class, 'ui-datepicker-other-month')) and @data-month='{month}' and @data-year='{year}']/a[text()='{day}']"
        day_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, day_xpath))
        )
        time.sleep(1)  # Thêm thời gian chờ
        day_element.click()

    except Exception as e:
        print(f"Lỗi trong quá trình xử lý: {e}")
#hàm tìm tổng số page
def check_and_click_page(driver):
    try:
        # Chờ đến khi phần tử có class 'j-bar-last' xuất hiện
        buttons = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, 'j-bar-last'))
        )
        time.sleep(3)
        
        # Tìm phần tử với thuộc tính 'disabled="disabled"' và click vào nó
        for button in buttons:
            if button.get_attribute('disabled') == None:
                driver.execute_script("arguments[0].click();", button)  # Sử dụng JavaScript để click
                print("Clicked on the disabled 'j-bar-last' button")
                break

        time.sleep(3)

        # Lấy giá trị sau dấu '/'
        span_text = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'j-bar-info'))
        ).text
        total_value = span_text.split('/')[-1].strip().strip(']')
        print(f"Total value after '/': {total_value}")
        
        return total_value

    except Exception as e:
        print(f"Lỗi trong quá trình xử lý: {e}")
        return None
def extract_header_data(driver):
    def remove_vietnamese_accents(text):
        # Chuyển đổi chuỗi thành dạng chuẩn NFD (Normalization Form D)
        normalized_text = unicodedata.normalize('NFD', text)
        # Loại bỏ các ký tự dấu (accent)
        without_accents = ''.join(c for c in normalized_text if unicodedata.category(c) != 'Mn')
        # Thay thế chữ Đ và đ thành D và d
        without_accents = without_accents.replace('Đ', 'D').replace('đ', 'd')
        return without_accents


    def remove_spaces(text):
        # Loại bỏ khoảng cách
        return re.sub(r'\s+', '', text)

    def process_text(text):
        text_without_accents = remove_vietnamese_accents(text)
        text_without_spaces = remove_spaces(text_without_accents)
        return text_without_spaces
    listbody_div_header = WebDriverWait(driver,10).until(
        EC.presence_of_element_located((By.ID,"lstMain-head"))
    )
    table_header = listbody_div_header.find_element(By.TAG_NAME,'table')
    tbody_header = table_header.find_elements(By.TAG_NAME,"tbody")
    rows = tbody_header[0].find_elements(By.TAG_NAME,"tr")
    data_header = [process_text(col.text) for row in rows for col in row.find_elements(By.TAG_NAME,"th")[1:]]
    data_header.insert(0, "STT")
    data_header.insert(1,"PAGE")
    return data_header
#hàm chọn tìm kiếm button theo ngày
def click_search_button(driver):
    global page 

    try: 
        Total = check_and_click_page(driver)
        if Total is not None:
           numberpage = int(Total) / 80
           numberpage_rounded = math.ceil(numberpage)
           print(f"Tổng số trang: {numberpage_rounded}")
        else:
           print("không thể lấy được tổng số trang lỗi")
        
        getdiv = driver.find_element(By.CLASS_NAME,'j-bar-warp')
        getclick = getdiv.find_element(By.CLASS_NAME,'j-bar-first')
        print(getclick.get_attribute('outerHTML'))
        if getclick.get_attribute('disabled') is None:
            getclick.click()
        data_header = extract_header_data(driver)
        print(data_header)
        dataheader = data_header
        totalpage = numberpage_rounded
        time.sleep(1)
        #lặp qua các page
           #
        if page == 0:
            # Không làm gì cả, tương đương với break khi i == 0
            pass
        elif page > 0:
            for i in range(page):
                if i > 0:  # Bỏ qua lần đầu tiên (i == 0)
                    click_next(driver)
                    time.sleep(1)
                    print(f"Iteration {i+1}")
                    numberpage = i+1
                    craw_data_cachup(driver) 
            
    except Exception as e:
        print(f"Lỗi trong quá trình xử lý: {e}")
#hàm login 
def login():
    try:
        chromedriver = "chromedriver.exe"
        urllogin = "http://192.168.0.65:8180/"
        username = "quyen.ngoq"
        password = "74777477"
        urldata = "http://192.168.0.65:8180/#menu=131&action=111"
        
        options = webdriver.ChromeOptions()
        # options.add_argument("--headless=new")
        # options.add_argument("--window-size=1920,1080")
        # options.add_argument("--disable-gpu")
        # options.add_argument("--no-sandbox")
        # options.add_argument("--disable-dev-shm-usage")
        # options.add_argument("--disable-extensions")
        # options.add_argument("--disable-plugins")
        # options.add_argument("--disable-software-rasterizer")
        # options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        service = Service(chromedriver)
        driver = webdriver.Chrome(service=service,options=options)
        #open the website
        driver.get(urllogin)
        driver.maximize_window()
        time.sleep(1)
        # Find and enter username
        username_input = driver.find_element(By.ID, "txtUsername")
        username_input.send_keys(username)

        # Find and enter password
        password_input = driver.find_element(By.ID, "txtPassword")
        password_input.send_keys(password)

        # Find and click the login button
        login_button = driver.find_element(By.ID, "btnLogin")
        login_button.click()

        # Wait for login to complete
        time.sleep(5)

        # Click save button
        save = driver.find_element(By.ID, "btnSave")
        save.click()
        time.sleep(3)
        
        #chọn danh mục ca chụp
        driver.get(urldata)
        time.sleep(2)
        set_date2(driver,"dbFrom",dateSelect)
        set_date2(driver,"dbTo",dateSelect)
        
        
        time.sleep(3)
        driver.quit()  
    except Exception as e:
        print(f"Lỗi trong quá trình thực thi chính:{e}")
    return driver

#hàm click next page
def click_next(driver):
    getdiv = driver.find_element(By.CLASS_NAME, 'j-bar-warp')
    getclick = getdiv.find_element(By.CLASS_NAME, "j-bar-next")
    print(getclick.get_attribute('outerHTML'))
    if getclick.get_attribute('disabled') == None:
       getclick.click()


#hàm đọc dữ liệu
def craw_data_cachup(driver):
    #lấy ngày trong file excel được chọn
    global dateSelect, urlFolder,dataheader
    #chuyển đổi ngày thành dd-mm-yyyy
    date = datetime.strptime(dateSelect,"%d/%m/%Y")
    formatted_date = date.strftime("%d-%m-%Y")
    csv_filenames = os.path.join(urlFolder,f"Cachup_data_{formatted_date}.csv")
    #đọc số lượng bản ghi hiện có trong csv file
    totalCsvNow = 0
    if os.path.exists(csv_filenames):
        with open(csv_filenames,'r',newline='',encoding='utf-8') as file:
            reader = csv.reader(file)
            totalCsvNow = sum(1 for row in reader) - 1
    else:
        print("===file csv không tồn tại===")
        with open(csv_filenames,'w',newline='',encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(dataheader)
            print(f"===khởi tạo file csv thành công tại:{csv_filenames}")
    print(f"===số lượng bản ghi hiện có trong file csv: {totalCsvNow}")

    def locate_table():
        listbody_div = WebDriverWait(driver,10).until(
            EC.presence_of_all_elements_located((By.ID,'lstMain-body'))
        )    
        table = listbody_div.find_element(By.TAG_NAME,'table')
        tbodys = table.find_elements(By.TAG_NAME,"tbody")
        return tbodys[1].find_elements(By.TAG_NAME,'tr')

    def get_row_data():
        rows = locate_table()
        data_array = []
        process_edit_data = []
        number = totalCsvNow + 1
        
        for row in rows[totalCsvNow:]:
            cols = row.find_elements(By.TAG_NAME,'td')
            data_row = [number,page]
            print(f"===đối tượng đầu tiên {number}===")
            number += 1
            numberIDBenhNhan = ""
            
            for col in cols[1:]:
                retry_count = 1
                while retry_count > 0:
                    try:
                         actions = ActionChains(driver)
                         actions.move_to_element(col).perform()
                         Text = WebDriverWait(driver,10).until(EC.visibility_of(col)).text
                         try:
                             icon_element = col.find_element(By.TAG_NAME,"i")
                             icon_html = icon_element.get_attribute('outerHTML')
                             text = "True"
                         except:
                             pass
                         print(f"nhánh {len(data_row)}:{text}")
                         if len(data_row) == 0:
                             numberIDBenhNhan = text
                         data_row.append(text)
                         
                         break
                    except StaleElementReferenceException:
                         retry_count -= 1
                         if retry_count == 0:
                             print(f"lỗi khi truy cập cột sau {3-retry_count} lần thử")
                             data_row.append("")
                         else:
                             print("Vào rồi nè")
                             time.sleep(0.5)
                             rows = locate_table()
                             time.sleep(0.5)
            print("===========================")
            #ghi dữ liệu tạm thời vào file csv
            print(f"có quyền ghi vào thư mục: {os.access(os.path.dirname(csv_filenames), os.W_OK)}")
            with open(csv_filenames,mode='a',newline='',encoding='utf-8') as file:
                writer = csv.DictWriter(file, fieldnames=dataheader)
                if file.tell() == 0:
                    writer.writeheader()
                writer.writerows([{dataheader[i]: data_row[i] for i in range(len(dataheader))}])
            data_array.append(data_row)   
        Csv_To_Excel(csv_filenames)
        return data_array,process_edit_data  
    
    data_array_all, process_edit_data = get_row_data()
    print("\n===== Hoàn tất hủy diệt dữ liệu =====")
    # def get_row_data():
    #     success = False
    #     while not success:
    #         number = 0
    #         if number > 20:
    #             login()
                
#hàm đọc dự liệu file excel


#hàm đọc dữ liệu file csv

#hàm main
def main():
    driver = login()
    click_search_button(driver)
    time.sleep(1)
    driver.quit()
    

#hàm chuyển đổi dữ liệu sang từ file csv sang excel
def Csv_To_Excel(csv_filename):
    try:
            df = pd.read_csv(csv_filename, encoding='utf-8-sig')
            output_filename = os.path.join(urlFolder, f"Patient_Data_{dateSelect}.xlsx")
            with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')

            # Định dạng file Excel
            wb = load_workbook(output_filename)
            ws = wb.active

            wb.save(output_filename)
            os.remove(csv_filename)  # Xóa file CSV sau khi chuyển sang Excel
            print("===kết thúc thu thập dữ liệu===")
    except Exception as e:
         print(f"lỗi chuyển đổi csv sang file excel: {e}")

#hàm thiết lập thông số app window
def select_folder():
    global urlFolder
    folder_path = filedialog.askdirectory(
        title="Select a folder"
    )
    if folder_path:
        print(folder_path)
        urlFolder = folder_path
        main()
    else:
        print("No folder selected.")
        

def select_Excel_File():
    file_path = filedialog.askopenfilename(
        title="Select an Excel File",
        filetypes=[("Excel files", "*.xlsx;*.xls")]
    )
    print(file_path)
    if file_path:
        Export_Excel_File_And_GetData(file_path)
    else:
        print("No file selected.")
        
def Export_Excel_File_And_GetData(file_path):
    global dateSelect
    if not file_path:
        print("Không có file nào được chọn.")
        return None
    # tìm tên file 
    file_name = os.path.basename(file_path)
    name_part, extension = os.path.splitext(file_name)
    try:
        date = datetime.strptime(name_part,"%d-%m-%Y")
        formatted_date = date.strftime("%d/%m/%Y")
        dateSelect = formatted_date
        
        # Đọc workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        print(formatted_date)
    except ValueError:
        print(name_part)
root = Tk()
root.title("Tkinter ComboBox Example")

#so 
numberget = [0]
# Đặt kích thước cho cửa sổ
window_width = 400
window_height = 300

# Lấy kích thước màn hình
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Tính toán vị trí để cửa sổ nằm giữa màn hình
position_x = (screen_width // 2) - (window_width // 2)
position_y = (screen_height // 2) - (window_height // 2)

# Đặt kích thước và vị trí cho cửa sổ
root.geometry(f'{window_width}x{window_height}+{position_x}+{position_y}')


date1 = ''
date2 = ''
    # Create a label to display instructions
label = ttk.Label(root, text="SIÊU PHẦN MỀM CRAW ĐÁ DỮ LIỆU ẢNH")
label.pack(pady=10)
# Create a label to display instructions
label = ttk.Label(root, text="Choose an option:")
label.pack(pady=10)


#get data json
file_button = ttk.Button(root, text="Select csv Excel", command=select_Excel_File, width=30)  # Corrected here
file_button.pack(pady=10)



button = ttk.Button(root, text="Get Data", command=select_folder, width=10)
button.pack(pady=10)

# Start the Tkinter event loop
root.mainloop()

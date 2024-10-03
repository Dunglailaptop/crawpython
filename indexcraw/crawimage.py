# import os
# import math
# import requests
# from selenium import webdriver
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.common.action_chains import ActionChains
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.support.ui import Select
# from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
# from PIL import Image
# from unidecode import unidecode 
# from tkinter import ttk, filedialog, Tk
# from datetime import datetime
# from tkinter import *
# from tkcalendar import DateEntry
# import time
# import csv
# import json
# import unicodedata
# import re
# import requests
# import io

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException,NoSuchElementException,ElementClickInterceptedException, TimeoutException
from concurrent.futures import ThreadPoolExecutor, as_completed
from unidecode import unidecode 
from tkinter import ttk, filedialog, Tk
from datetime import datetime
from tkinter import *
from tkcalendar import DateEntry
from tkinter.filedialog import askopenfilename
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import threading
import pandas as pd
import openpyxl
import pandas as pd
import math
import time
import csv
import json
import unicodedata
import re
import requests
import io
import os
from PIL import Image
import numpy as np


#urlfolder lưu đường dẫn folder khi ghi nhận xong
urlFolder = ""
#urlFileExcel lưu đường dẫn file Excel
urlFileExcel = ""
#ngày chọn 
dateSelect = ''

dateSelect1 = ''
dateSelect2 = ''
# page = 2
urlFileCSV = ''

def call_api_import_size(json_file_path, url, chunk_size=100):
    with open(json_file_path, 'r', encoding='utf-8', errors='ignore') as file:
        data = json.load(file)

    headers = {
        "Content-Type": "application/json"
    }

    def import_chunk(chunk):
        try:
            response = requests.post(url, json=chunk, headers=headers)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"Error importing chunk: {e}")
            return None

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = []
        for i in range(0, len(data), chunk_size):
            chunk = data[i:i+chunk_size]
            print(f"Sending chunk {i//chunk_size + 1} (size: {len(chunk)})")
            futures.append(executor.submit(import_chunk, chunk))

        for future in as_completed(futures):
            result = future.result()
            if result:
                print(json.dumps(result, indent=2))

    print("All data chunks have been processed.")

def call_api_import(datapost,urls):
    url = urls
    print(datapost)
    
    dataapitest = [{
        "ID": "56181185",
        "Ma": "KTC01",
        "Ten": "T\u00fai Ti\u1ec3u C\u1ea7u [40ml]",
        "Tenviettat": "T\u00fai Ti\u1ec3u C\u1ea7u [40ml]",
        "Loai": "M\u00e1u",
        "Thetich": "40 ml",
        "Gia": "449,167 \u20ab",
        "Ngaytao": "20/05/2022",
        "Uutien": "0"
    }]
    
    headers = {
        "Content-Type": "application/json"
    }

    response = requests.post(url, json=datapost,headers=headers)

    if response.status_code == 200:
        data = response.json()
        print(data)
        return data
    else:
        print(f"Yêu cầu thất bại với mã lỗi: {response.status_code}")
        Message = f"Yêu cầu thất bại với mã lỗi: {response.status_code}"
        return Message

def remove_vietnamese_accents(text):
    # Chuyển đổi chuỗi thành dạng chuẩn NFD (Normalization Form D)
    normalized_text = unicodedata.normalize('NFD', text)
    # Loại bỏ các ký tự dấu (accent)
    without_accents = ''.join(c for c in normalized_text if unicodedata.category(c) != 'Mn')
    # Thay thế chữ Đ và đ thành D và d
    without_accents = without_accents.replace('Đ', 'D').replace('đ', 'd')
    return without_accents


def remove_spaces(text):
    # Loại bỏ khoảng cách
    return re.sub(r'\s+', '', text)

def process_text(text):
    text_without_accents = remove_vietnamese_accents(text)
    text_without_spaces = remove_spaces(text_without_accents)
    return text_without_spaces

def next_action(driver,area_data_url):
    global dateSelect1,dateSelect2
# chọn phân trang tính toán tổng số page
    driver.get(area_data_url)
    time.sleep(2)
    set_date2(driver, "dbFrom","01/01/2023")
    set_date2(driver, "dbTo", "31/01/2023")
    time.sleep(5)

        # Tìm phần tử có class "btns"
    btns_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, "btns"))
    )

# Tìm phần tử có id "btnSearch" bên trong phần tử "btns"
    search_button = btns_element.find_element(By.ID, "btnSearch")

# Click vào nút tìm kiếm
    search_button.click()



# Function to initialize and log into the system
def login(type):
    global dateSelect, urlFolder
    try:
        login_url = "http://192.168.0.65:8180/"
        area_data_url = ""
        username = "quyen.ngoq"
        password = "74777477"
        area_data_url = "http://192.168.0.65:8180/#menu=131&action=111"  
            
        download_dir = os.path.join(urlFolder)
        os.makedirs(download_dir, exist_ok=True)

        # Initialize ChromeDriver
        options = webdriver.ChromeOptions()
        
        options.add_experimental_option("prefs", {
            "download.default_directory": urlFolder,
            "profile.default_content_setting_values.automatic_downloads": 1,
            "download.prompt_for_download": False,
            "profile.default_content_settings.popups": 0,
            "safebrowsing.enabled": "false",
            "safebrowsing.disable_download_protection": True
         })
        prefs = {"credentials_enable_service": False,
                 "profile.password_manager_enabled": False}
        options.add_experimental_option("prefs", prefs)
        options.add_argument(f"--unsafely-treat-insecure-origin-as-secure={login_url}")
        options.add_argument(f"--unsafely-treat-insecure-origin-as-secure={area_data_url}")
        options.add_argument("--headless=new")  # Chạy trình duyệt ở chế độ ẩn
        options.add_argument("--window-size=1920x1080")  # Kích thước cửa sổ mặc định
        #thêm của claude hướng dẫn
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument('--safebrowsing-disable-download-protection')
     
        #========================
        driver = webdriver.Chrome(options=options)

        # Mở website
        driver.get(login_url)
        driver.maximize_window()
        # Sử dụng WebDriverWait để đảm bảo phần tử đã sẵn sàng
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "txtUsername")))
        
        # Nhập username
        username_input = driver.find_element(By.ID, "txtUsername")
        username_input.send_keys(username)

        # Nhập password
        password_input = driver.find_element(By.ID, "txtPassword")
        password_input.send_keys(password)

        # Nhấp nút đăng nhập
        login_button = driver.find_element(By.ID, "btnLogin")
        login_button.click()

        # Chờ đợi cho đến khi trang hoàn tất load sau khi đăng nhập
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "btnSave")))
        
        # Nhấp nút save (nếu cần)
        save_button = driver.find_element(By.ID, 'btnSave')
        save_button.click()

        # Xử lý theo loại yêu cầu
        if int(type) == 1:
            next_action(driver, area_data_url)
        else:
            url = "http://192.168.0.65:8180/#menu=29&action=168"
            driver.get(url)
            print("Thực hiện chỉ định")

        print("=== Đăng nhập hoàn tất ===")
    except Exception as e:
        print(f"Lỗi trong quá trình thực thi: {e}")

    return driver


def select_date(driver, month_value, year_value):
    try:
           # Chờ đợi thẻ <select> của năm xuất hiện
        year_select_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'ui-datepicker-year'))
        )
        # Tạo đối tượng Select và chọn giá trị năm
        year_select = Select(year_select_element)
        year_select.select_by_value(str(year_value))
        time.sleep(1)  # Thêm thời gian chờ
        # Chờ đợi thẻ <select> của tháng xuất hiện
        month_select_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'ui-datepicker-month'))
        )
        # Tạo đối tượng Select và chọn giá trị tháng
        month_select = Select(month_select_element)
        print("tháng đã chọn vào hệ thống:" + str(month_value))
        month_select.select_by_value(str(month_value))
        time.sleep(5)  # Thêm thời gian chờ

     

        print(f"Đã chọn tháng {month_value + 1} và năm {year_value}")

    except Exception as e:
        print(f"Lỗi trong quá trình xử lý: {e}")

def set_date(driver, element_id, date_value):
    try:
        # Chờ đến khi phần tử có ID xuất hiện
        searchIdDateTime = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, element_id))
        )

        # Tìm thẻ <span> với class 'input-group-addon' bên trong phần tử
        findspan = searchIdDateTime.find_element(By.CLASS_NAME, 'input-group-addon')

        # Click vào thẻ <span> để mở lịch
        findspan.click()
        time.sleep(2)  # Thêm thời gian chờ để đảm bảo lịch được mở

        # Tách date_value thành ngày, tháng và năm
        day, month, year = date_value.split('/')
        print(month)
        # Gọi hàm select_date để chọn tháng và năm
        select_date(driver, int(month) - 1, year)  # Chuyển đổi tháng về dạng số và trừ 1 vì tháng trong datepicker bắt đầu từ 0

        # Chọn ngày
        day_xpath = f"//td[not(contains(@class, 'ui-datepicker-other-month')) and @data-month='{month}' and @data-year='{year}']/a[text()='{day}']"
        day_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, day_xpath))
        )
        time.sleep(1)  # Thêm thời gian chờ
        day_element.click()

    except Exception as e:
        print(f"Lỗi trong quá trình xử lý: {e}")

        


#check disabled và lấy tổng số phần tử
def check_and_click_page(driver):
    try:
        time.sleep(15)
        wait_for_fe_load(driver)
        # Chờ đến khi phần tử có class 'j-bar-last' xuất hiện
        buttons = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, 'j-bar-last'))
        )
        time.sleep(5)
        
        # Tìm phần tử với thuộc tính 'disabled="disabled"' và click vào nó
        for button in buttons:
            if button.get_attribute('disabled') == None:
                driver.execute_script("arguments[0].click();", button)  # Sử dụng JavaScript để click
                print("Clicked on the disabled 'j-bar-last' button")
                break

        time.sleep(5)

        # Lấy giá trị sau dấu '/'
        span_text = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'j-bar-info'))
        ).text
        total_value = span_text.split('/')[-1].strip().strip(']')
        print(f"Total value after '/': {total_value}")
        
        return total_value

    except Exception as e:
        print(f"Lỗi trong quá trình xử lý: {e}")
        return None
#call back
def check_and_click_page_callback(driver):
    try:
        # Chờ đến khi phần tử có class 'j-bar-last' xuất hiện
        buttons = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, 'j-bar-first'))
        )
        time.sleep(3)
        
        # Tìm phần tử với thuộc tính 'disabled="disabled"' và click vào nó
        for button in buttons:
            if button.get_attribute('disabled') == None:
                driver.execute_script("arguments[0].click();", button)  # Sử dụng JavaScript để click
                print("Clicked on the disabled 'j-bar-last' button")
                break

        time.sleep(3)
    except Exception as e:
        print(f"Lỗi trong quá trình xử lý: {e}")
        return None
 #ngay 2
def set_date2(driver, element_id, date_value):
    try:
        # Chờ đến khi phần tử có ID xuất hiện
        searchIdDateTime = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, element_id))
        )

        # Tìm thẻ <span> với class 'input-group-addon' bên trong phần tử
        findspan = searchIdDateTime.find_element(By.CLASS_NAME, 'input-group-addon')

        # Click vào thẻ <span> để mở lịch
        findspan.click()
        time.sleep(2)  # Thêm thời gian chờ để đảm bảo lịch được mở

        # Tách date_value thành ngày, tháng và năm
        day, month, year = date_value.split('/')
        day = int(day)
        month = int(month) - 1  # Chuyển đổi tháng về dạng số và trừ 1 vì tháng trong datepicker bắt đầu từ 0
        year = int(year)

        # Gọi hàm select_date để chọn tháng và năm
        select_date(driver, month, year)

        # Chọn ngày
        day_xpath = f"//td[not(contains(@class, 'ui-datepicker-other-month')) and @data-month='{month}' and @data-year='{year}']/a[text()='{day}']"
        day_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, day_xpath))
        )
        time.sleep(1)  # Thêm thời gian chờ
        day_element.click()

    except Exception as e:
        print(f"Lỗi trong quá trình xử lý: {e}")

def select_date(driver, month, year):
    try:
        # Chọn năm
        year_select = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'ui-datepicker-year'))
        )
        year_select.click()
        year_option = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//option[@value='{year}']"))
        )
        year_option.click()

        # Chọn tháng
        month_select = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'ui-datepicker-month'))
        )
        month_select.click()
        month_option = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//option[@value='{month}']"))
        )
        month_option.click()

    except Exception as e:
        print(f"Lỗi khi chọn tháng và năm: {e}")
##       
def click_next(driver):
    try:
        # Đợi cho đến khi phần tử j-bar-warp xuất hiện
        wait = WebDriverWait(driver, 10)
        getdiv = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'j-bar-warp')))
        
        # Tìm nút "next" trong phần tử j-bar-warp
        getclick = getdiv.find_element(By.CLASS_NAME, "j-bar-next")
        print(getclick.get_attribute('outerHTML'))
        
        # Kiểm tra xem nút có bị vô hiệu hóa không
        if getclick.get_attribute('disabled') is None:
            # Đợi cho đến khi nút có thể được nhấp
            wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "j-bar-next")))
            getclick.click()
            # Đợi cho đến khi trang mới được tải
            wait.until(EC.staleness_of(getdiv))
        else:
            print("Nút 'next' đã bị vô hiệu hóa")
            return False
    except (StaleElementReferenceException, TimeoutException) as e:
        print(f"Lỗi khi nhấp nút 'next': {e}")
        return False
    return True
#hàm lấy đọc dữ liệu file excel lấy stt và page cuối cùng
def read_excel_last_element():
    global urlFileExcel
    try:
        # Đọc file Excel
        df = pd.read_excel(urlFileExcel)
        
        # Kiểm tra xem có ít nhất hai cột không
        if len(df.columns) < 2:
            return {"STT": 0, "Page": 0}
        
        # Lấy hai cột đầu tiên
        stt = df.iloc[:, 0]
        page = df.iloc[:, 1]
        
        # Kiểm tra xem có dữ liệu không
        if len(stt) == 0 or len(page) == 0:
            return {"STT": 0, "Page": 0}
        
        # Lấy phần tử cuối cùng
        last_element = {
            "STT": stt.iloc[-1],
            "Page": page.iloc[-1]
        }
        
        return last_element
    except Exception:
        # Trả về một từ điển với giá trị mặc định nếu có bất kỳ lỗi nào xảy ra
        return {"STT": 0, "Page": 0}
def select_csv_file():
    global urlFileCSV
    # Ẩn cửa sổ chính của tkinter
    Tk().withdraw()

    # Mở hộp thoại để chọn file
    file_path = askopenfilename(filetypes=[("CSV files", "*.csv")])
    
    # Kiểm tra xem có chọn file hay không
    if not file_path:
        print("KO TÌM THẤY FILE")
    urlFileCSV = file_path 
def read_csv_last_element():
    global urlFileCSV
    try:
        # Đọc file CSV
        df = pd.read_csv(urlFileCSV)
        
        # Kiểm tra xem có ít nhất hai cột không
        if len(df.columns) < 2:
            return {"STT": 0, "Page": 0}
        
        # Lấy hai cột đầu tiên
        stt = df.iloc[:, 0]
        page = df.iloc[:, 1]
        
        # Kiểm tra xem có dữ liệu không
        if len(stt) == 0 and len(page) == 0:
            return {"STT": 0, "Page": 0}
        
        # Lấy phần tử cuối cùng
        last_element = {
            "STT": stt.iloc[-1],
            "Page": page.iloc[-1]
        }
        
        return last_element
    except Exception:
        # Trả về một từ điển với giá trị mặc định nếu có bất kỳ lỗi nào xảy ra
        return {"STT": 0, "Page": 0}

#hàm lấy total và tổng số page
def get_total_and_page(driver):
    print("===bắt đầu lấy tổng số phần tử và page===")
    Total = 0
    page = 0
    try:    
        
        Total = check_and_click_page(driver)
        if Total is not None:
            numberpage = int(Total) / 80
            numberpage_rounded = math.ceil(numberpage)
            print(f"Tổng số page ghi nhận là: {numberpage_rounded}")
            page = numberpage_rounded
        else:
            print("lỗi ghi nhận số page")
        
        getdiv = driver.find_element(By.CLASS_NAME, 'j-bar-warp')
        getclick = getdiv.find_element(By.CLASS_NAME, "j-bar-first")
        # print(getclick.get_attribute('outerHTML'))
        if getclick.get_attribute('disabled') is None:
           getclick.click()
        time.sleep(10)
    except Exception as e:
        print("====lỗi hàm get total====")
        if driver:
            driver.quit()
        login_again()
    return Total,page

def click_search_button(driver):
    global page
    try:    
        Total = check_and_click_page(driver)
        if Total is not None:
            numberpage = int(Total) / 80
            numberpage_rounded = math.ceil(numberpage)
            print(f"Number of pages (rounded up): {numberpage_rounded}")
        else:
            print("Total is None, unable to calculate the number of pages.")
        
        getdiv = driver.find_element(By.CLASS_NAME, 'j-bar-warp')
        getclick = getdiv.find_element(By.CLASS_NAME, "j-bar-first")
        print(getclick.get_attribute('outerHTML'))
        if getclick.get_attribute('disabled') is None:
           getclick.click()
        time.sleep(10)
       
        
      
        # iterations = numberpage_rounded
        # success = False
        # number_now = 0
        # #doc file csv 
            
        # current_page = page
        # csv_filenames = ""
        # urlCsvFile = ""
        # formatted_date = ""
        # while current_page < iterations:
        #     time.sleep(1)
        #     data_header = []
        #     data_header = extract_header_data(driver)
        #     print(data_header)
        #     if current_page == 0:
        #         # Xử lý trang đầu tiên
        #         csv_filenames,urlCsvFile,formatted_date,success,number_now = extract_and_save_table_data_loads_cachup(driver, current_page, data_header, urlFolder)
            
        #         Csv_To_Excel(csv_filenames,urlCsvFile,formatted_date)
        #     else:
        #         click_next(driver)
        #         print(f"page số {current_page + 1}")
        #         csv_filenames,urlCsvFile,formatted_date, success,number_now = extract_and_save_table_data_loads_cachup(driver, current_page, data_header, urlFolder)
            
        #         Csv_To_Excel(csv_filenames,urlCsvFile,formatted_date)
        #     current_page += 1
          
                 
                   
         
        
    except Exception as e:
        print(f"Lỗi trong quá trình xử lý: {e}")



   



# Function to extract header data
def extract_header_data(driver):
    listbody_div_header = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "lstMain-head"))
    )
    table_header = listbody_div_header.find_element(By.TAG_NAME, 'table')
    tbody_header = table_header.find_elements(By.TAG_NAME, "tbody")
    rows = tbody_header[0].find_elements(By.TAG_NAME, "tr")
    data_header = [process_text(col.text) for row in rows for col in row.find_elements(By.TAG_NAME, "th")[1:]]
    return data_header


# Function to extract table data and save to CSV and JSON
def clean_price(price_str):
    return float(re.sub(r'[^\d.]', '', price_str))

def parse_date(date_str):
    return datetime.strptime(date_str, "%d/%m/%Y").isoformat()


#new function
def capture_image(driver, download_path, file_name, numberId):
    try:
        # Chụp ảnh toàn màn hình
        png = driver.get_screenshot_as_png()
        
        # Sử dụng PIL để mở ảnh từ dữ liệu PNG
        im = Image.open(io.BytesIO(png))

        # Tạo đường dẫn đầy đủ cho file
        folder_path = os.path.join(download_path, str(numberId))
        full_path = os.path.join(folder_path, file_name)

        # Tạo thư mục nếu chưa tồn tại
        os.makedirs(folder_path, exist_ok=True)

        # Lưu ảnh
        im.save(full_path)

        print(f"Đã lưu ảnh: {full_path}")
        print("success")
    except Exception as e:
        print("failed -", str(e))
#lấy dữ liệu 
#lấy dữ liệu file csv
def getdatacsv(driver):
    global dateSelect,urlFileCSV
    data_header = extract_header_data(driver)
    data_header.insert(0,"STT")
    data_header.insert(1,"PAGE")
    data_header.insert(2,"KETQUACHIDINHKHAM")
    # date = datetime.strptime(dateSelect,"%d/%m/%Y")
    # formatted_date = date.strftime("%d-%m-%Y")
    csv_filenames = os.path.join(urlFileCSV)
    # Đọc số lượng bản ghi hiện có trong file CSV
    existing_records = 0
    if os.path.exists(csv_filenames):
        with open(csv_filenames, 'r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            existing_records = sum(1 for row in reader) - 1  # Trừ 1 để bỏ qua header
    else: 
        print("===không tồn tại file CSV===")
        with open(csv_filenames, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(data_header)
            print(f"Đã tạo thành công csv tai: {csv_filenames}")
    print(f"Số bản ghi hiện có: {existing_records}")
    return existing_records ,data_header,csv_filenames
#LẤY PAGE VÀ STT
def locate_table(driver):
    listbody_div = WebDriverWait(driver, 0.1).until(
        EC.presence_of_element_located((By.ID, 'lstMain-body'))
    )
    table = listbody_div.find_element(By.TAG_NAME, 'table')
    tbodys = table.find_elements(By.TAG_NAME, "tbody")
    return tbodys[1].find_elements(By.TAG_NAME, 'tr')

def get_series_data(number, numberId,driver):
    global urlFolder
    series_data = []
    checkButton = False
    # Tìm tất cả các phần tử có class 'series-item'
    series_items = WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.series-item'))
    )
    numberSeries_item = 0
    for series_item in series_items:
        numberSeries_item += 1
        series_item.click()
        time.sleep(0.5)
        
        # Lấy giá trị của thẻ 'ins-length'
        
        image_count = int(WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.ins-length'))
        ).text)
        print("in số lượng ảnh:" + str(image_count))

        # Chụp ảnh và lưu vào thư mục
        if image_count > 1:
            for i in range(image_count):
                capture_image(driver, urlFolder,
                            f"saved_image_{number}_frame_{i}_{numberSeries_item}.png", numberId)
                time.sleep(3)
                # Nhấn nút chuyển sang ảnh tiếp theo
                next_button = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '.paging .fa-chevron-right'))
                )
                next_button.click()
                time.sleep(0.5)
        else:
            capture_image(driver, urlFolder,
                        f"saved_image_{number}_frame_0_{numberSeries_item}.png", numberId)
        
        series_data.append({
            "image_count": image_count
        })

    return series_data
def wait_for_fe_load(driver, timeout=30):
    script = """
    return new Promise((resolve) => {
        if (typeof window.FELoad === 'function') {
            let originalFELoad = window.FELoad;
            window.FELoad = function() {
                originalFELoad.apply(this, arguments);
                resolve(true);
            };
        } else {
            // Nếu FELoad không tồn tại, giả định rằng nó đã được gọi
            resolve(true);
        }
        
        // Timeout sau một khoảng thời gian nếu FELoad không được gọi
        setTimeout(() => resolve(false), arguments[0]);
    });
    """
    
    try:
        is_loaded = WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script(script, timeout * 1000)
        )
        if is_loaded:
            print("FELoad đã được thực thi!")
            return True
        else:
            print("Timeout khi đợi FELoad thực thi")
            return False
    except TimeoutException:
        print("Timeout khi đợi FELoad thực thi")
        return False
def getDownload(driver, numberIdPatient):
    try:
        # Đợi và nhấn nút Download
        div_check_button_download = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, 'btnDownload'))
        )
        div_check_button_download.click()

        # print(f"benh nhan:{numberIdPatient}")

        # Đợi cho loading mask biến mất
        WebDriverWait(driver, 30).until(
            EC.invisibility_of_element_located((By.ID, "mask"))
        )

        # Đợi và nhấn nút Save
        button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "btnSave"))
        )
        button.click()
        # Sử dụng JavaScript để nhấn nút thay vì click trực tiếp
        # driver.execute_script("arguments[0].click();", button)

        # Đợi cho đến khi nút Save biến mất
        WebDriverWait(driver, 30).until_not(
            EC.presence_of_element_located((By.ID, "btnSave"))
        )

        # urldown = "C:\\Users\\ndung\\Downloads"
        # getfiledownloadnow = os.path.join(urldown)
        # print(f"Đường dẫn thư mục tải xuống: {getfiledownloadnow}")

        # # Đợi cho file tải xuống hoàn tất
        # wait_for_download_complete(urldown)

    except Exception as e:
        print(f"Lỗi trong quá trình thực thi: {str(e)}")
        # Có thể thêm logic retry ở đây nếu cần

def wait_for_download_complete(download_path, timeout=300):
    start_time = time.time()
    while time.time() - start_time < timeout:
        time.sleep(1)
        files = os.listdir(download_path)
        if any(file.endswith('.crdownload') for file in files):
            print("Đang tải xuống...")
        else:
            print("Tải xuống hoàn tất!")
            return True
    print("Hết thời gian chờ tải xuống")
    return False
   
   


def getData_Image(cols, number, numberId,driver):
        # Get the first column data (assuming it's the ID or unique identifier)
        # cols[0].click()
        time.sleep(0.5)  # Đợi để trang load

        div_Menu = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'menuCrud'))
        )
        Button_showImage = div_Menu.find_element(By.ID, 'btnWebViewer')
        Button_showImage.click()
        time.sleep(3)
        div_Zoom = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".layout-menu-right button[title='Toàn màn hình']"))
        )
        div_Zoom.click()
        time.sleep(0.5)

        series_data = get_series_data(number, numberId,driver)
        # print("tong so luong anh lay dc:"+series_data)

        time.sleep(0.5)
        div_ViewImage = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".layout-menu-right button[title='Thu nhỏ (ALT-M)']"))
        )
        time.sleep(0.5)
        div_ViewImage.click()
        time.sleep(3)
        # cols[0].click()
        # time.sleep(3)
# craw ca chụp
def extract_and_save_table_data_loads_cachup(driver,Stt,numberRun,page):
    global dateSelect, urlFolder
    # Đọc số lượng bản ghi hiện có trong file CSV
    numberrunget = 0
    #
    existing_records,data_header,csv_filenames = getdatacsv(driver)
    print("=======lay du lieu ca chup========")
    time.sleep(0.5)

    

    def click_element(element):
        try:
            # Đợi cho đến khi phần tử có thể click được
            element = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, element.get_attribute('xpath')))
            )
            # Scroll đến phần tử
            driver.execute_script("arguments[0].scrollIntoView(true);", element)
            # Click vào phần tử
            element.click()
        except TimeoutException:
            # Nếu không thể click, thử dùng JavaScript
            driver.execute_script("arguments[0].click();", element)    
    
    def get_many_image(MaBenhNhan):
        headers = ["MaBenhNhan"]
        csvFilenameImageToLong = os.path.join(urlFolder,"DS_BenhNhan_Nhieu_Anh")
        if os.path.exists(csvFilenameImageToLong):
            with open(csvFilenameImageToLong, 'a', newline='', encoding='utf-8') as file:
                writer = csv.DictWriter(file, fieldnames=headers)
                if file.tell() == 0:  # If the file is empty, write the header
                    writer.writeheader()
                writer.writerow({"MaBenhNhan": MaBenhNhan})
        else:
            with open(csvFilenameImageToLong, 'w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(headers)
                writer.writerow([MaBenhNhan])
            print(f"Đã tạo thành công file csv của thư mục nhiều ảnh tại: {csvFilenameImageToLong}")
        print("đã ghi xong bệnh nhân nhiều ảnh")
            
    
    def get_row_data(page):
        rows = locate_table(driver)
        data_array = []
        process_edit_data = []
      
        
        #logic
        total_press = existing_records
        start_index = total_press % 80
        number = total_press + 1
        ketquakhamchidinh = ""
        number_now = 0
        for row in rows[start_index:]:
            number_now += 1
            cols = row.find_elements(By.TAG_NAME, 'td')
            data_row = [number,page,ketquakhamchidinh]
            print(f"===đối tượng đầu tiên {number}===")
            number += 1
            numberIDBenhNhan = ""
            if number_now == 81:
                return data_array, process_edit_data, number_now
            for col in cols[1:]:
                retry_count = 1
               
                while retry_count > 0:
                    try:
                        actions = ActionChains(driver)
                        actions.move_to_element(col).perform()
                        text = WebDriverWait(driver, 0.1).until(EC.visibility_of(col)).text
                        try:
                            icon_element = col.find_element(By.TAG_NAME, "i")
                            icon_html = icon_element.get_attribute('outerHTML')
                            text = "True"
                        except:
                            pass
                        print(f"nhánh {len(data_row)}: {text}")
                        if len(data_row) == 3:   
                           numberIDBenhNhan = text
                        if len(data_row) == 8:
                            print(f"mã bệnh nhân là: {numberIDBenhNhan}")
                            valueImage = text
                            if valueImage.isdigit():  # Kiểm tra xem valueImage có phải là số không
                                cols[0].click()
                                getDownload(driver)
                                print(f"thẻ html của click: {cols[0].get_attribute('outerHTML')}")
                                # actions = ActionChains(driver)
                                # actions.move_to_element(cols[0]).perform()
                                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                                # Hoặc cuộn đến phần tử cụ thể
                                element = cols[0]  # hoặc phần tử bạn muốn cuộn đến
                                driver.execute_script("arguments[0].scrollIntoView(true);", element)
                                time.sleep(0.2)
                                cols[0].click()
                              
                                # if int(valueImage) <= 2000:  # Thay đổi từ 5 thành 3
                                #     cols[0].click()
                                #     getData_Image(cols, number, numberIDBenhNhan,driver)
                                #     print(f"thẻ html của click: {cols[0].get_attribute('outerHTML')}")
                                #     # actions = ActionChains(driver)
                                #     # actions.move_to_element(cols[0]).perform()
                                #     driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                                #     # Hoặc cuộn đến phần tử cụ thể
                                #     element = cols[0]  # hoặc phần tử bạn muốn cuộn đến
                                #     driver.execute_script("arguments[0].scrollIntoView(true);", element)
                                #     time.sleep(5)
                                #     cols[0].click()
                                # else:
                                #     print("ghi nhận lại bệnh nhân có nhiều ảnh từ 3 trở lên")
                                #     print(f"Bỏ qua bệnh nhân {numberIDBenhNhan} vì có {valueImage} hình ảnh")
                                #     cols[0].click()
                                #     getData_Image(cols, number, numberIDBenhNhan,driver)
                                #     print(f"thẻ html của click: {cols[0].get_attribute('outerHTML')}")
                                #     # actions = ActionChains(driver)
                                #     # actions.move_to_element(cols[0]).perform()
                                #     driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                                #     # Hoặc cuộn đến phần tử cụ thể
                                #     element = cols[0]  # hoặc phần tử bạn muốn cuộn đến
                                #     driver.execute_script("arguments[0].scrollIntoView(true);", element)
                                #     time.sleep(5)
                                #     cols[0].click()
                                #     # get_many_image(numberIDBenhNhan)
                                #     return data_array, process_edit_data, number_now
                            else:
                                print(f"Giá trị không hợp lệ cho số lượng hình ảnh: {valueImage}")
                        data_row.append(text)
                        
                        
                        break
                    except StaleElementReferenceException:
                        retry_count -= 1
                        if retry_count == 0:
                            print(f"Lỗi khi truy cập cột sau {3 - retry_count} lần thử")
                            data_row.append("")
                        else:
                            print("Vào rồi nè")
                            time.sleep(0.5)
                            rows = locate_table(driver)
                            time.sleep(0.5)
            print("===============================")
           
            # original_window = driver.current_window_handle
            # patient_id = data_row[3]
            # loaixetnghiem = data_row[6]
            # if driver:
            #     driver.execute_script("window.open('');")
            #     driver.switch_to.window(driver.window_handles[-1])
            #     url = {
            #         "US": "http://192.168.0.65:8180/#menu=29&action=168",
            #         "CR": "http://192.168.0.65:8180/#menu=23&action=130",
            #         "MR": "http://192.168.0.65:8180/#menu=20&action=159",
            #         "CT": "http://192.168.0.65:8180/#menu=17&action=133"
            #     }.get(loaixetnghiem, "")
            #     driver.get(url)
            #     print(f"Thực hiện chỉ định cho bệnh nhân ID: {patient_id}")
            #     try:
            #         datas = get_patient_data(patient_id, driver)
            #         print(f"Retrieved data: {datas}")
            #         data_row[2] = datas
            #     except Exception as e:
            #         print(f"Lỗi khi xử lý bệnh nhân {patient_id}")
            #     finally:
            #         driver.close()
            #         driver.switch_to.window(original_window)
            #     print(f"Đã hoàn thành xử lý cho bệnh nhân ID: {patient_id}")

            # print("Đã quay về trang web ban đầu và hoàn thành tất cả bệnh nhân")
                
            # # Ghi dữ liệu tạm thời vào file CSV
            print(f"Có quyền ghi vào thư mục: {os.access(os.path.dirname(csv_filenames), os.W_OK)}")
            with open(csv_filenames, mode='a', newline='', encoding='utf-8') as file:
                writer = csv.DictWriter(file, fieldnames=data_header)
                if file.tell() == 0:  # Nếu file rỗng, ghi header
                    writer.writeheader()
                writer.writerows([{data_header[i]: data_row[i] for i in range(len(data_header))}])
            data_array.append(data_row)
            
            # Csv_To_Excel(csv_filenames)
            # if number <= 10:   
            #    getData_Image(cols,number,numberIDBenhNhan)
        return data_array, process_edit_data, number_now
    
    data_array_all, process_edit_data, numberrunget = get_row_data(page)
    print(data_array_all)
    
    return numberrunget,csv_filenames 
#xử lý ngoại lệ
def login_again(max_retries=3):
    for attempt in range(max_retries):
        try:
            driver = login(1)
            
            # Thực hiện các bước đăng nhập
            return driver
        except WebDriverException as e:
            print(f"Lỗi khi đăng nhập (lần thử {attempt + 1}): {e}")
            if attempt < max_retries - 1:
                time.sleep(5)  # Chờ trước khi thử lại
            else:
                print("===chạy lại hàm main đã thử 2 lần có vẻ hệ thống reset sau 300s===")
                raise
                # raise  # Nếu đã thử hết số lần, ném ngoại lệ
# hàm cãi tiến
def extract_and_save_table_data_loads_cachup_permon(driver, Stt, numberRun, page):
    global dateSelect, urlFolder
    existing_records, data_header, csv_filenames = getdatacsv(driver)
    print("=======lay du lieu ca chup========")

    def get_row_data(page):
        rows = locate_table(driver)
        number = existing_records + 1
        startindex = existing_records % 80
        processed_count = 0

        for row in rows[startindex:]:
            cols = row.find_elements(By.TAG_NAME, 'td')
            data_row = [number, page, ""]  # Khởi tạo với các giá trị mặc định

            numberIDBenhNhan = cols[1].text  # Lấy mã bệnh nhân
            tenbenhnhan = cols[2].text
            print(f"====bệnh nhân số {number} với mã: {numberIDBenhNhan}====")
            print(f"==>tên bệnh nhân:{tenbenhnhan}")
            print("===========================")
            for col in cols[1:]:
                try:
                    actions = ActionChains(driver)
                    actions.move_to_element(col).perform()
                    text = WebDriverWait(driver, 0.1).until(EC.visibility_of(col)).text 
                    if len(data_row) == 8:  # Xử lý cột hình ảnh
                        if text.isdigit():
                            cols[0].click()
                            getDownload(driver,numberIDBenhNhan)
                            driver.execute_script("arguments[0].scrollIntoView(true);", cols[0])
                            time.sleep(0.5)
                            cols[0].click()
                    data_row.append(text)
                except StaleElementReferenceException:
                    data_row.append("")

            # Ghi dữ liệu ngay lập tức vào file CSV
            with open(csv_filenames, mode='a', newline='', encoding='utf-8') as file:
                writer = csv.DictWriter(file, fieldnames=data_header)
                if file.tell() == 0:
                    writer.writeheader()
                writer.writerow({data_header[i]: data_row[i] for i in range(len(data_header))})
            #Csv_To_Excel(csv_filenames)
            number += 1
            processed_count += 1
            if processed_count >= 20:
                break

        return processed_count

    numberrunget = get_row_data(page)
    return numberrunget, csv_filenames
# Main function to run the script
def main():
    global urlFolder, dateSelect
    max_try = 5
    retry_count = 0

    while retry_count < max_try:
        try:
           
            total = 0
            totalpage = 0
            driver = None
            csvfilenew = ""

            # Đăng nhập lần đầu để lấy giá trị total
            driver = login_again()
            total, totalpage = get_total_and_page(driver)
            print(f"+=>hoàn tất ghi nhận số page là:{totalpage}")
            print(f"+=>hoàn tất ghi nhận tổng số phần tử cần lấy là: {total}")
            
            driver.quit()

            # Lấy dữ liệu từ file Excel
            # result = read_excel_last_element()
            readCSVFILE = read_csv_last_element()
            print(f"{readCSVFILE["STT"]} va {readCSVFILE["Page"]}")
            Stt = int(readCSVFILE["STT"])
            page = int(readCSVFILE["Page"])
            if Stt == 0 and page == 0:
                Stt = 0
                page = 0
            print(f"+=>ta có Stt:{Stt}, page:{page}")

            numberget = Stt
            items_per_page = 80

            while numberget < int(total):
                print(f"====đang chạy: {numberget}/{total}====")
                page = min(numberget // items_per_page, totalpage - 1)
                
                driver = login_again()
                for _ in range(page):
                    if not click_next(driver):
                        print("Không thể tiếp tục điều hướng")
                        break
                time.sleep(1)

                items_extracted, csvfilenew = extract_and_save_table_data_loads_cachup_permon(driver, Stt, numberget, page)
                # Csv_To_Excel(csvfilenew)
                
                numberget += items_extracted
                Stt += items_extracted
                
                driver.quit()
            print("===Kết thúc thu thập dữ liệu===")
            print("Hoàn thành quá trình lấy dữ liệu.")
            break

        except Exception as e:
            retry_count += 1
            print(f"Lỗi trong quá trình thực thi chính: {e} thử chạy lại {retry_count}/{max_try}")
            if driver:
                driver.quit()
            if retry_count < max_try:
                print(f"Đang thử lại sau 10 giây...")
                time.sleep(10)
            else:
                print("Đã vượt quá số lần thử lại tối đa. Kết thúc chương trình.")
                break



def on_button_click():
    print("Button clicked!")
    

def select_folder():
    global urlFolder
    folder_path = filedialog.askdirectory(
        title="Select a folder"
    )
    if folder_path:
        print(folder_path)
        urlFolder = folder_path
        main()
    else:
        print("No folder selected.")
        

def select_Excel_File():
    file_path = filedialog.askopenfilename(
        title="Select an Excel File",
        filetypes=[("Excel files", "*.xlsx;*.xls")]
    )
    print(file_path)
    if file_path:
        Export_Excel_File_And_GetData(file_path)
    else:
        print("No file selected.")
   
#hàm truy vết toàn bộ ảnh còn lại
def get_data_image_final():
    global dateSelect, urlFolder
    folder_path = filedialog.askdirectory(
        title="Select a folder"
    )
    if folder_path:
        print(folder_path)
        urlFolder = folder_path
      
    else:
        print("No folder selected.")
    csvFilenameImageToLong = os.path.join(urlFolder, "DS_BenhNhan_Nhieu_Anh")
    namefolder = os.path.basename(urlFolder)
    print(f"ten folder: {namefolder}")
    date = namefolder.split("_")[-1]
    formatdate = datetime.strptime(date, "%d-%m-%Y")
    convert = formatdate.strftime("%d/%m/%Y")
    dateSelect = convert
    patient_ids = []
    if os.path.exists(csvFilenameImageToLong):
        print(f"File CSV đã được tìm thấy tại: {csvFilenameImageToLong}")
        with open(csvFilenameImageToLong, 'r', newline='', encoding='utf-8') as file:
            csv_reader = csv.DictReader(file)
            for row in csv_reader:
                patient_ids.append(row['MaBenhNhan'])
        
        print(f"Đã đọc {len(patient_ids)} mã bệnh nhân từ file CSV.")
        print("===thực thi truy vết từng cái===")
        
        #login lần đầu để setup khung giờ
        driver = login_again()
        numbers = False
        numbersget = 0
        try:
            for id in patient_ids:
                if numbers:
                    if driver:
                        driver.quit()    
                    driver = login_again()
                
                searchIdPatient = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "txtPatientId"))
                )
                searchIdPatient.clear()
                searchIdPatient.send_keys(id)

                btns_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "btns"))
                )
                search_button = btns_element.find_element(By.ID, "btnSearch")
                search_button.click()
                time.sleep(3)
                rows = locate_table(driver)
                for row in rows:
                    cols = row.find_elements(By.TAG_NAME, "td")
                    if len(cols) > 6:
                        image_count = cols[6].text
                        if image_count.isdigit() and int(image_count) > 3:
                            print(f"tìm thấy bệnh nhân nhiều ảnh: {image_count}")
                            print("===thực hiện lấy ảnh===")
                            cols[0].click()
                            getData_Image(cols, numbersget, int(cols[1].text), driver)
                            print(f"Lấy xong mã bệnh nhân:{cols[1].text}")
                            numbers = True
                            # Removed break to process all matching patients
                if not numbers:
                    print(f"Không tìm thấy bệnh nhân có nhiều hơn 3 ảnh cho ID: {id}")
                numbersget += 1
        finally:
            if driver:
                driver.quit()
        


#đọc file excel 
def Export_Excel_File_And_GetData(file_path):
    global dateSelect,urlFileExcel
    if not file_path:
        print("Không có file nào được chọn.")
        return None
    # tìm tên file 
    file_name = os.path.basename(file_path)
    name_part, extension = os.path.splitext(file_name)
    try:
        date = datetime.strptime(name_part,"%d-%m-%Y")
        formatted_date = date.strftime("%d/%m/%Y")
        dateSelect = formatted_date
        urlFileExcel = file_path
        print(f"đường dẫn:{file_path}")
        # Đọc workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        print(formatted_date)
    except ValueError:
        print(name_part)
#ghi toàn bộ
def Csv_To_Excel_Overite(csv_filename):
    global urlFileExcel
    try:
        output_filename = os.path.basename(urlFileExcel)
        
        # Đọc file CSV vào DataFrame
        df = pd.read_csv(csv_filename, encoding='utf-8-sig')
        
        # Tạo ExcelWriter object
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            # Ghi DataFrame vào Excel
            df.to_excel(writer, sheet_name='Sheet1', index=False)
        
        # Xóa file CSV gốc
      
        print("===Kết thúc thu thập dữ liệu===")
        return True

    except Exception as e:
        print(f"Lỗi chuyển đổi CSV sang file Excel: {e}")
        return False
#ghi từng dòng
def Csv_To_Excel(csv_filename):
    global urlFileExcel
    try:
        if not os.path.exists(urlFileExcel):
            print(f"File Excel không tồn tại: {urlFileExcel}")
            return False

        # Mở workbook hiện có
        wb = load_workbook(urlFileExcel)
        
        # Chọn sheet đầu tiên hoặc tạo mới nếu không có
        if len(wb.sheetnames) > 0:
            ws = wb.active
        else:
            ws = wb.create_sheet(title="Sheet1")

        # Đọc file CSV và ghi từng dòng vào Excel
        with open(csv_filename, 'r', encoding='utf-8-sig') as csvfile:
            csv_reader = csv.reader(csvfile)
            headers = next(csv_reader)  # Đọc header
            
            # Tìm dòng trống đầu tiên trong sheet
            first_empty_row = 1
            while ws.cell(row=first_empty_row, column=1).value is not None:
                first_empty_row += 1
            
            # Ghi headers nếu sheet trống
            if first_empty_row == 1:
                for col, header in enumerate(headers, start=1):
                    ws.cell(row=first_empty_row, column=col, value=header)
                first_empty_row += 1
            
            # Ghi dữ liệu
            for row_idx, row in enumerate(csv_reader, start=first_empty_row):
                for col, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col, value=value)
                
                # Lưu workbook sau mỗi 1000 dòng để giảm sử dụng bộ nhớ
                if row_idx % 1000 == 0:
                    wb.save(urlFileExcel)

        # Lưu lần cuối
        wb.save(urlFileExcel)
        print(f"Dữ liệu đã được thêm vào file Excel: {urlFileExcel}")
        return True

    except Exception as e:
        print(f"Lỗi chuyển đổi CSV sang file Excel: {e}")
        return False
    
def choose_csv_file():    
    file_path = filedialog.askopenfilename(
        title="Chọn file CSV",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
    )
    
    if file_path:
        print(f"File đã chọn: {file_path}")
        return file_path
    else:
        print("Không có file nào được chọn.")
        return None


def get_patient_data(patient_id, driver):
    global dateSelect
    data_Array = []
    def get_data_in_chidinh(number_row):
        div_process = driver.find_element(By.ID, "processEdit")
        
        # Use a dictionary to store all the data
        data = {
            "description": div_process.find_element(By.ID, "description").text,
            "ketluan": div_process.find_element(By.ID, "conclusion").text,
            "KyThuat": div_process.find_element(By.ID, "technical").text,
            "DeNghi": div_process.find_element(By.ID, "comments").text,
            "ChuanDoan": div_process.find_element(By.ID, "diagnosis").text,
            "NoiChiDinh": div_process.find_element(By.ID, "department").get_attribute('value'),
            "BacSiChiDinh": div_process.find_element(By.ID, "provider").get_attribute('value'),
            "MauKetQua": div_process.find_element(By.ID, "temp").find_element(By.CSS_SELECTOR, "input").get_attribute('value'),
            "ThietBi": div_process.find_element(By.ID, "device").find_element(By.CSS_SELECTOR, "input").get_attribute('value'),
            "MaAnhDICOM": div_process.find_element(By.ID, "accessionId").get_attribute('value'),
            "NoiThucHien": div_process.find_element(By.ID, "performedDepartment").find_element(By.CSS_SELECTOR, "input").get_attribute('value'),
            "BSDocKetQua": div_process.find_element(By.ID, "performedProvider").find_element(By.CSS_SELECTOR, "input").get_attribute('value'),
            "KTVThucHien": div_process.find_element(By.ID, "technician").find_element(By.CSS_SELECTOR, "input").get_attribute('value'),
            "MaPhieuChiDinh": div_process.find_element(By.ID, "resultCode").get_attribute('value')
        }
        # print(f"====chidinh-số thứ tự:-dòng:{number_row}-Mabenhnhan:{patient_id}====")
        # print(data)
        # for key, value in data.items():
        #     print(f"+==={key}: {value}")
        # print("==========================")
        return data
    def find_and_click_row(driver, number_row):
        max_attempts = 3
        for attempt in range(max_attempts):
            try:
                # Wait for the table to be present
                table = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "table-striped"))
                )
                
                # Find all rows
                rows = table.find_elements(By.CLASS_NAME, "j-listitem")
                
                if number_row <= len(rows):
                    row = rows[number_row - 1]
                    
                    # Scroll the row into view
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'nearest'});", row)
                    
                    # Wait for the row to be clickable
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, row.get_attribute('id'))))
                    
                    # Click the row
                    row.click()
                    return True
                else:
                    print(f"Row {number_row} not found")
                    return False
            except (StaleElementReferenceException, ElementClickInterceptedException, TimeoutException, NoSuchElementException) as e:
                print(f"Attempt {attempt + 1} failed for row {number_row}: {e}")
                if attempt == max_attempts - 1:
                    print(f"Failed to click row {number_row} after {max_attempts} attempts")
                    return False
                driver.refresh()  # Refresh the page before retrying
                time.sleep(2)  # Wait for the page to reload

   
    def setup():
        set_date2(driver, "dbFrom", dateSelect)
        set_date2(driver, "dbTo", dateSelect)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "txtPatient")))
        patient_input = driver.find_element(By.ID, "txtPatient")
        patient_input.clear()
        patient_input.send_keys(patient_id)
        
        search_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "btnSearch")))
        search_button.click()
        
        # Wait for the table to be present
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "table-striped")))

    try:
        setup()
        time.sleep(1.5)
        # Get the total number of rows
        table = driver.find_element(By.CLASS_NAME, "table-striped")
        rows = table.find_elements(By.CLASS_NAME, "j-listitem")
        total_rows = len(rows)
        print(f"tổng số hồ sơ tìm thấy là:{total_rows}")
        for number_row in range(1, total_rows + 1):
            if find_and_click_row(driver, number_row):
                try:
                    # Wait for the patient info to be present
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "service-info"))
                    )
                    
                    # Get patient info
                    patient_info = {
                        "patient_id": driver.find_element(By.CLASS_NAME, "patient-id").text,
                        "order_date": driver.find_element(By.CLASS_NAME, "order-date").text,
                        "order_num": driver.find_element(By.CLASS_NAME, "order-num").text,
                        "patient_name": driver.find_element(By.CLASS_NAME, "patient-name").text,
                        "service_count": driver.find_element(By.CLASS_NAME, "service-count").text,
                        "service_name": driver.find_element(By.CLASS_NAME, "service-name").text
                    }
                    # print(f"Patient Info: {patient_info}")
                    if patient_info["patient_id"] == patient_id:
                        time.sleep(1)
                        data = get_data_in_chidinh(number_row)
                        data_Array.append(data)
                    else:
                        print("===rỗng không có kết quả===")
                except (TimeoutException, NoSuchElementException) as e:
                    print(f"Error getting patient info for row {number_row}: {e}")
            else:
                print(f"Skipping row {number_row} due to click failure")
    except TimeoutException:
        print("Trang web mất quá nhiều thời gian để phản hồi")
        return None
    return data_Array


def threading():
    global dateSelect1, dateSelect2
    def my_function1(name, delay):
        print(f"Luồng {name} bắt đầu")
        time.sleep(delay)
        dateSelect1 = '01/02/2023'
        dateSelect2 = '31/02/2023'
        main()
        print(f"Luồng {name} kết thúc sau {delay} giây")
    def my_function2(name, delay):
        print(f"Luồng {name} bắt đầu")
        time.sleep(delay)
        main()
        print(f"Luồng {name} kết thúc sau {delay} giây")

    # Tạo hai luồng
    thread1 = threading.Thread(target=my_function, args=("1", 2))
    thread2 = threading.Thread(target=my_function, args=("2", 4))

    # Bắt đầu chạy các luồng
    thread1.start()
    thread2.start()

    # Đợi cho đến khi cả hai luồng kết thúc
    thread1.join()
    thread2.join()

    print("Tất cả các luồng đã hoàn thành")
   
   
  

root = Tk()
root.title("Tkinter ComboBox Example")

#so 
numberget = [0]
# Đặt kích thước cho cửa sổ
window_width = 400
window_height = 300

# Lấy kích thước màn hình
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Tính toán vị trí để cửa sổ nằm giữa màn hình
position_x = (screen_width // 2) - (window_width // 2)
position_y = (screen_height // 2) - (window_height // 2)

# Đặt kích thước và vị trí cho cửa sổ
root.geometry(f'{window_width}x{window_height}+{position_x}+{position_y}')


date1 = ''
date2 = ''
    # Create a label to display instructions
label = ttk.Label(root, text="SIÊU PHẦN MỀM CRAW ĐÁ DỮ LIỆU ẢNH")
label.pack(pady=10)
# Create a label to display instructions
label = ttk.Label(root, text="Choose an option:")
label.pack(pady=10)


#get data json
file_button = ttk.Button(root, text="Select csv Excel", command=select_csv_file, width=30)  # Corrected here
file_button.pack(pady=10)



button = ttk.Button(root, text="Get Data", command=select_folder, width=10)
button.pack(pady=10)

button = ttk.Button(root, text="Get data final many image", command=get_data_image_final, width=10)
button.pack(pady=10)

# button = ttk.Button(root, text="lấy chỉ định khám ",command=doc_chidinh_CLS,width=10)
# button.pack(pady=10)
# Start the Tkinter event loop
root.mainloop()



# # Run the main function
# if __name__ == "__main__":
#     khoitaoapp()
